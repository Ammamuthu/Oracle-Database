import paramiko
import getpass
import time
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# SSH Connection
def connect_to_server(hostname, username, password, max_retries=1):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    retry = 0
    while retry <= max_retries:
        try:
            ssh.connect(hostname, username=username, password=password, timeout=10)
            return ssh
        except Exception as e:
            print(f"❌ Connection failed to {hostname}: {e}")
            retry += 1
            time.sleep(2)
    return None

# Run remote command
def run_command(ssh, command):
    try:
        stdin, stdout, stderr = ssh.exec_command(command)
        return stdout.read().decode()
    except Exception as e:
        print(f"❌ Command error: {e}")
        return ""

# Discover OPatch paths
def find_opatch_paths(ssh):
    find_cmd = "find /u01/app -type f -name opatch 2>/dev/null"
    output = run_command(ssh, find_cmd)
    opatch_paths = []
    for line in output.strip().splitlines():
        if '/OPatch/opatch' in line:
            opatch_paths.append(line.strip())
    return opatch_paths

# Extract patch info + date
def parse_weblogic_patches_with_dates(lspatches_output, lsinventory_output):
    patches = []

    patch_dates = {}
    for line in lsinventory_output.splitlines():
        match = re.search(r'Patch\s+(\d+)\s+: applied on (.+)', line)
        if match:
            patch_id = match.group(1)
            raw_date = match.group(2).strip()
            try:
                parts = raw_date.split()
                if len(parts) == 6:
                    date_str = " ".join(parts[:4] + parts[5:])
                else:
                    date_str = raw_date
                parsed_date = datetime.strptime(date_str, "%a %b %d %H:%M:%S %Y")
                patch_dates[patch_id] = parsed_date.strftime("%Y-%m-%d %H:%M:%S")
            except:
                patch_dates[patch_id] = raw_date

    for line in lspatches_output.strip().splitlines():
        match = re.match(r"(\d+);(.+)", line.strip())
        if match:
            patch_id = match.group(1)
            description = match.group(2).strip()
            version_match = re.search(r'(\d+\.\d+\.\d+\.\d+\.\d+)', description)
            release_version = version_match.group(1) if version_match else ""
            patch_date = patch_dates.get(patch_id, "")
            patches.append({
                "patch_id": patch_id,
                "description": description,
                "release_version": release_version,
                "applied_date": patch_date
            })
    return patches

# Export Excel
def export_to_excel(data_dict, filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "WebLogic Patch Info"

    headers = ["Hostname", "Patch ID", "Description", "Release Version", "Applied Date"]

    # Styles
    header_font = Font(bold=True, color="000000")  # Black text
    header_fill = PatternFill(start_color="66FFB2", end_color="66FFB2", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")

    # Header row
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    row_num = 3  # Leave one row gap after headers

    for hostname, patches in data_dict.items():
        for patch in patches:
            ws.cell(row=row_num, column=1, value=hostname).alignment = center_align
            ws.cell(row=row_num, column=2, value=patch["patch_id"]).alignment = center_align
            ws.cell(row=row_num, column=3, value=patch["description"]).alignment = center_align
            ws.cell(row=row_num, column=4, value=patch["release_version"]).alignment = center_align
            ws.cell(row=row_num, column=5, value=patch["applied_date"]).alignment = center_align
            row_num += 1
        row_num += 1  # 🔽 Extra space between servers

    ws.freeze_panes = "A3"
    wb.save(filename)
    print(f"\n✅ Excel file created: {filename}")

# Main function
def main():
    servers_input = input("Enter hostnames/IPs (comma-separated or file path): ").strip()
    if os.path.isfile(servers_input):
        with open(servers_input) as f:
            servers = [line.strip() for line in f if line.strip()]
    else:
        servers = [s.strip() for s in servers_input.split(",")]

    # Try default credentials
    username = "imsadmin"
    password = "Ecssupport09"

    patch_data = {}

    for server in servers:
        print(f"\n🔗 Connecting to {server} with default credentials...")
        ssh = connect_to_server(server, username, password)
        if not ssh:
            print("⚠️ Default credentials failed.")
            username = input("Enter SSH username: ").strip()
            password = getpass.getpass("Enter SSH password: ")
            ssh = connect_to_server(server, username, password)

        if ssh:
            print(f"✅ Connected to {server}")
            patch_data[server] = []
            opatch_paths = find_opatch_paths(ssh)

            if not opatch_paths:
                print(f"⚠️ No OPatch found on {server}")
                ssh.close()
                continue

            for opatch_path in opatch_paths:
                base_path = os.path.dirname(os.path.dirname(opatch_path))
                print(f"📦 Found WebLogic Home: {base_path}")

                lspatches_cmd = f"{opatch_path} lspatches"
                lsinv_cmd = f"{opatch_path} lsinventory"

                lspatches_output = run_command(ssh, lspatches_cmd)
                lsinv_output = run_command(ssh, lsinv_cmd)

                if lspatches_output:
                    patches = parse_weblogic_patches_with_dates(lspatches_output, lsinv_output)
                    patch_data[server].extend(patches)
                    print(f"✅ {len(patches)} patches found in {base_path}")
                else:
                    print(f"⚠️ No lspatches output from {base_path}")

            ssh.close()
        else:
            print(f"❌ Could not connect to {server}")
            patch_data[server] = []

    if patch_data:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"weblogic_patches_{timestamp}.xlsx"
        export_to_excel(patch_data, filename)
    else:
        print("❌ No patch data collected.")

if __name__ == "__main__":
    main()
